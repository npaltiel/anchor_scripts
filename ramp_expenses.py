import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

current = datetime.now() - timedelta(days=30)

expenses_df = pd.read_csv(
    f"C:\\Users\\nochum.paltiel\\OneDrive - Anchor Home Health care\\Documents\\Ramp Expenses\\Ramp Expenses {current.month}-{current.year}.csv")

categories_df = pd.read_csv(
    "C:\\Users\\nochum.paltiel\\OneDrive - Anchor Home Health care\\Documents\\Ramp Expenses\\Category Dropdown.csv")
category_options = [option for option in categories_df['Category']]
expenses_df = expenses_df[
    ['Transaction Time', 'Amount', 'User', 'User Email', 'Merchant Name', 'Merchant Description', 'Card Last 4',
     'Accounting Category', 'Accounting Class']]

users = expenses_df[['User', 'User Email']].drop_duplicates().reset_index(drop=True)

user_list = {users['User'][i]: expenses_df[expenses_df['User Email'] == users['User Email'][i]] for i in
             range(len(users['User Email']))}


def manage_excel_file(file_path, sheet_name, df):
    """
    Write a DataFrame to an Excel file.

    Parameters:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to write to.
        df (pd.DataFrame): The DataFrame to write.
    """
    if os.path.exists(file_path):
        # File exists: Load the workbook and write to a new sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            workbook = load_workbook(file_path)
            if sheet_name in workbook.sheetnames:
                print(f"Sheet '{sheet_name}' in '{user}' already exists")
            else:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"DataFrame written to existing file '{user}' in sheet '{sheet_name}'.")
    else:
        # File does not exist: Create a new workbook and write to the first sheet
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Created new workbook '{user}' and wrote DataFrame to sheet '{sheet_name}'.")

        # Add dropdowns for Confirmed Category and Approved columns
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Get the starting column for new columns
    start_col = sheet.max_column + 1

    # Define headers for new columns
    confirmed_category_col = get_column_letter(start_col)
    approved_col = get_column_letter(start_col + 1)

    sheet[f"{confirmed_category_col}1"] = "Confirmed Category"
    sheet[f"{approved_col}1"] = "Approved"

    # Define dropdowns
    category_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(category_options)}"'
    )
    approved_validation = DataValidation(
        type="list",
        formula1='"Yes,No"'
    )

    # Apply dropdowns to all rows in the new columns
    for row in range(2, sheet.max_row + 1):  # Starting from row 2 (after the header)
        category_cell = f"{confirmed_category_col}{row}"
        approved_cell = f"{approved_col}{row}"

        sheet[category_cell].value = None
        sheet[approved_cell].value = None

        category_validation.add(sheet[category_cell])
        approved_validation.add(sheet[approved_cell])

    # Add validations to the sheet
    sheet.add_data_validation(category_validation)
    sheet.add_data_validation(approved_validation)

    # Left-align the first row
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal='left')

    # Resize columns to fit the content
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2  # Add padding for better fit

    # Save the workbook
    workbook.save(file_path)
    print(f"DataFrame written with dropdowns to '{sheet_name}' in '{user}'")


for user in user_list:
    file = f"C:\\Users\\nochum.paltiel\\OneDrive - Anchor Home Health care\\Documents\\Ramp Expenses\\Ramp Expenses - {user}.xlsx"
    manage_excel_file(file, f'{current.strftime('%b')}-{current.year}', user_list[user])
