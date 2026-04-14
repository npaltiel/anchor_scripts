import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# FILE PATHS
# ============================================================

contracts_path = r"C:\Users\nochum.paltiel\OneDrive - Anchor Home Health care\Documents\General Information\Contract Lookup.csv"
raw_path = r"C:\Users\nochum.paltiel\OneDrive - Anchor Home Health care\Documents\Underutilization Report\Files\Authorizations_Under_Utilized.csv"
report_path_prev = r"C:\Users\nochum.paltiel\OneDrive - Anchor Home Health care\Documents\Underutilization Report\Underutilized Weekly Reports\Underutilization Report - 03-04-2026.xlsx"
report_path_out = r"C:\Users\nochum.paltiel\OneDrive - Anchor Home Health care\Documents\Underutilization Report\Underutilized Weekly Reports\Underutilization Report - 03-30-2026.xlsx"
hist_path = r"C:\Users\nochum.paltiel\OneDrive - Anchor Home Health care\Documents\Underutilization Report\Files\Historical Table.xlsx"

# ============================================================
# COLUMN WIDTH RULES (NO HUGE COLUMNS)
# ============================================================

MAX_COL_WIDTH = 30  # absolute cap
NOTES_MAX_WIDTH = 25  # Notes / Previous Notes columns cap
DEFAULT_COL_WIDTH = 15  # normal cap
MIN_COL_WIDTH = 10


def autosize_columns_capped(ws):
    """
    - Notes/Previous Notes get a bit wider but capped
    - Everything else capped tighter
    - Report Notes is treated as a normal column (so it won't explode on Raw Data)
    """
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        header = ws[f"{col_letter}3"].value  # headers are on row 3 (startrow=2)

        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        if header in ("Previous Notes", "Notes"):
            width = min(max_len + 3, NOTES_MAX_WIDTH)
        else:
            width = min(max_len + 2, DEFAULT_COL_WIDTH)

        ws.column_dimensions[col_letter].width = max(
            MIN_COL_WIDTH,
            min(width, MAX_COL_WIDTH)
        )


# ============================================================
# 1. READ & PREP RAW DATA
# ============================================================

raw_df = pd.read_csv(raw_path, skiprows=3)

raw_df.columns = [
    "Auth Type", "Row Num", "AdmissionID", "Patient Name", "Auth Period",
    "ScheduledTime", "Contract", "Coordinator", "Auth Number",
    "Report Notes", "Visit Date", "Authorized Hours", "Assigned Hours",
    "Discipline", "Service Code", "NurseAssigned", "PatientStatus", "ProgramCode"
]

raw_df = raw_df[raw_df["Auth Type"] == "Daily"]

raw_df["Visit Date"] = pd.to_datetime(raw_df["Visit Date"], errors="coerce")
raw_df["Assigned Hours"] = pd.to_numeric(raw_df["Assigned Hours"], errors="coerce").fillna(0)

run_date = raw_df["Visit Date"].max().date()

# 1. Coordinator Only
raw_df["Coordinator Only"] = raw_df["Coordinator"].astype(str).str.split(",", n=1).str[0]

# 2. Team
tmp = raw_df["Coordinator Only"].astype(str).str.replace("PCA ", "", regex=False)
tmp2 = tmp.str.split(" ", n=1).str[0]
raw_df["Team"] = tmp2.str.split("-", n=1).str[0]

# Merge contract lookup
contracts_df = pd.read_csv(contracts_path)
contracts_df.columns = contracts_df.columns.str.strip()

raw_df = raw_df.merge(
    contracts_df[["ContractName", "ContractType"]],
    left_on="Contract",
    right_on="ContractName",
    how="left"
).drop(columns=["ContractName"])

# Enforce T4 = NHTD only (Team exists now)
raw_df = raw_df[
    (raw_df["Team"] != "T4") |
    (raw_df["ContractType"] == "NHTD")
    ]

# 3. Underutilized
raw_df["Amount Underutilized"] = raw_df["Authorized Hours"] - raw_df["Assigned Hours"]

# 4. Overall daily
raw_df["Amount Underutilized Overall daily"] = (
    raw_df.groupby(["AdmissionID", "Visit Date"])["Amount Underutilized"].transform("sum")
)

# 5. Day of the Week
raw_df["Day of the Week"] = raw_df["Visit Date"].dt.day_name()

# 6. Consistency count
raw_df["Consistent"] = (
    raw_df.groupby(["AdmissionID", "Day of the Week", "Contract"])["Visit Date"]
    .transform(lambda x: x.nunique())
)

# 7. Minimum Underutilized
raw_df["Amount Consistently Underutilized"] = (
    raw_df.groupby(["AdmissionID", "Day of the Week", "Contract"])
    ["Amount Underutilized Overall daily"].transform("min")
)

# 8. Recused?
recused_list = [
    "PCA T3 Hold or Refusing",
    "PCA T3 Pending Onboarding CG",
    "PCA T6 pending CG",
    "PA - New admissions",
    "PCA T3 Case Management Review"
]
raw_df["Recused?"] = raw_df["Coordinator Only"].isin(recused_list)

# ============================================================
# 2. READ PRIOR REPORT & NORMALIZE OUTCOME (AND PRESERVE LEGACY BEHAVIOR)
# ============================================================

team_tabs = ["T1", "T2", "T3", "T4", "T5", "T6", "PA"]
frames = []

for tab in team_tabs:
    df_prev = pd.read_excel(report_path_prev, sheet_name=tab, skiprows=2)
    frames.append(df_prev)

teams_df = pd.concat(frames, ignore_index=True)
teams_df = teams_df.dropna(subset=["AdmissionID"])

# Clean blanks
if "Notes" in teams_df.columns:
    teams_df["Notes"] = teams_df["Notes"].replace("", pd.NA)
else:
    teams_df["Notes"] = pd.NA

# Preserve prior behavior: New Notes = Notes if filled else Previous Notes
if "Previous Notes" not in teams_df.columns:
    teams_df["Previous Notes"] = pd.NA

teams_df["Previous Notes"] = teams_df["Notes"].fillna(teams_df["Previous Notes"])

# Outcome: prefer Outcome col; else one-time map Approved boolean; else blank
if "Outcome" in teams_df.columns:
    teams_df["Outcome"] = teams_df["Outcome"].replace("", pd.NA)
elif "Approved" in teams_df.columns:
    # Only TRUE maps to Approved; FALSE/blank treated as blank (i.e., no action)
    teams_df["Outcome"] = teams_df["Approved"].apply(lambda x: "Approved" if x is True else pd.NA)
else:
    teams_df["Outcome"] = pd.NA

teams_df["Outcome"] = teams_df["Outcome"].fillna("").astype(str).str.strip()

# ============================================================
# 3. READ / EXTEND HISTORY & APPEND SNAPSHOT
# ============================================================

try:
    hist_df = pd.read_excel(hist_path, sheet_name="Data")
except FileNotFoundError:
    hist_df = pd.DataFrame()

required_cols = [
    "Date", "Coordinator", "AdmissionID", "Patient Name", "Day of the Week",
    "Minimum Underutilized", "Previous Notes",
    "Previously Approved", "Previously Approved Hours",
    "Previously Ignored", "Previously Ignored Hours",
    "Previously Fixed", "Previously Fixed Hours"
]

for c in required_cols:
    if c not in hist_df.columns:
        hist_df[c] = pd.NA

# Normalize history types
hist_df["Date"] = pd.to_datetime(hist_df["Date"], errors="coerce").dt.date

for b in ["Previously Approved", "Previously Ignored", "Previously Fixed"]:
    hist_df[b] = hist_df[b].fillna(False).astype(bool)

for h in ["Previously Approved Hours", "Previously Ignored Hours", "Previously Fixed Hours"]:
    hist_df[h] = pd.to_numeric(hist_df[h], errors="coerce").fillna(0)

# Build snapshot to append
teams_df["Date"] = run_date

teams_df["Previously Approved"] = teams_df["Outcome"].eq("Approved")
teams_df["Previously Ignored"] = teams_df["Outcome"].eq("Ignore")
teams_df["Previously Fixed"] = teams_df["Outcome"].eq("Fixed")

teams_df["Previously Approved Hours"] = teams_df["Minimum Underutilized"].where(teams_df["Previously Approved"], 0)
teams_df["Previously Ignored Hours"] = teams_df["Minimum Underutilized"].where(teams_df["Previously Ignored"], 0)
teams_df["Previously Fixed Hours"] = teams_df["Minimum Underutilized"].where(teams_df["Previously Fixed"], 0)

# Only append the required history columns
hist_append = teams_df[required_cols]

updated_history = pd.concat([hist_df, hist_append], ignore_index=True)
updated_history["Date"] = pd.to_datetime(updated_history["Date"], errors="coerce").dt.date

# Write history WITHOUT deleting other sheets:
# - openpyxl: replace Data sheet content by rewriting workbook while preserving others
try:
    wb_hist = load_workbook(hist_path)
    # Remove existing Data sheet if present
    if "Data" in wb_hist.sheetnames:
        ws_old = wb_hist["Data"]
        wb_hist.remove(ws_old)
    wb_hist.save(hist_path)
except FileNotFoundError:
    pass

with pd.ExcelWriter(hist_path, engine="openpyxl", mode="a") as writer:
    updated_history.to_excel(writer, sheet_name="Data", index=False)

# ============================================================
# 4. MERGE LATEST HISTORY INTO RAW
# ============================================================

hist_latest = (
    updated_history.sort_values("Date", ascending=False)
    .drop_duplicates(subset=["AdmissionID", "Day of the Week"], keep="first")
)

raw_df = raw_df.merge(
    hist_latest[[
        "AdmissionID", "Day of the Week", "Previous Notes",
        "Previously Approved", "Previously Approved Hours",
        "Previously Ignored", "Previously Ignored Hours",
        "Previously Fixed", "Previously Fixed Hours"
    ]],
    on=["AdmissionID", "Day of the Week"],
    how="left"
)

# Fill defaults
raw_df["Previous Notes"] = raw_df["Previous Notes"].fillna("")
for b in ["Previously Approved", "Previously Ignored", "Previously Fixed"]:
    raw_df[b] = raw_df[b].fillna(False).astype(bool)
for h in ["Previously Approved Hours", "Previously Ignored Hours", "Previously Fixed Hours"]:
    raw_df[h] = pd.to_numeric(raw_df[h], errors="coerce").fillna(0)


# ============================================================
# 5. BUILD TEAM TABLES (KEEP ALL PREVIOUS COLUMNS + NEW ONES)
# ============================================================

def build_team_table(df, team):
    # Exclude only if Approved/Ignore AND current hours <= stored threshold (same as old behavior extended)
    suppressed = (
            (df["Previously Approved"] & (df["Amount Consistently Underutilized"] <= df["Previously Approved Hours"])) |
            (df["Previously Ignored"] & (df["Amount Consistently Underutilized"] <= df["Previously Ignored Hours"]))
    )

    mask = (
            (df["Consistent"] == 3) &
            (df["Team"] == team) &
            (~df["Recused?"]) &
            (~suppressed)
    )

    out = df.loc[mask, [
        "Coordinator Only",
        "AdmissionID",
        "Patient Name",
        "Day of the Week",
        "Amount Consistently Underutilized",
        "Previously Approved",
        "Previously Approved Hours",
        "Previously Ignored",
        "Previously Ignored Hours",
        "Previous Notes",
    ]]

    return (
        out.rename(columns={
            "Coordinator Only": "Coordinator",
            "Amount Consistently Underutilized": "Minimum Underutilized"
        })
        .drop_duplicates()
        .sort_values(["Coordinator", "AdmissionID"])
    )


tables = {t: build_team_table(raw_df, t) for t in team_tabs}


# ============================================================
# 6. WRITE REPORT
# ============================================================

def write_sheet(writer, df, name):
    df = df.copy()

    # Add Notes/Outcome only on team tabs (same behavior as before, just renamed Approved->Outcome)
    if name != "Raw Data":
        df["Notes"] = ""
        df["Outcome"] = ""

    df.to_excel(writer, sheet_name=name, index=False, startrow=2)


with pd.ExcelWriter(report_path_out, engine="openpyxl", mode="w") as writer:
    write_sheet(writer, raw_df, "Raw Data")
    for t in team_tabs:
        write_sheet(writer, tables[t], t)

# ============================================================
# 7. POST-FORMAT: COLUMN WIDTHS + OUTCOME DROPDOWN
# ============================================================

wb = load_workbook(report_path_out)

for ws in wb.worksheets:
    # Capped autosize (prevents Report Notes from exploding)
    autosize_columns_capped(ws)

    # Outcome dropdown on team sheets
    if ws.title != "Raw Data":
        outcome_col = None
        for cell in ws[3]:  # header row
            if cell.value == "Outcome":
                outcome_col = cell.column_letter
                break

        if outcome_col:
            dv = DataValidation(type="list", formula1='"Approved,Fixed,Ignore"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{outcome_col}4:{outcome_col}{ws.max_row}")

wb.save(report_path_out)
